"""
Build docs/meetings/2026-05-25-notifications-decision-matrix.xlsx.

Matches the format of the existing review-window-decision-matrix.xlsx
exactly:
  - 5 sheets (one per module + Cross-cutting Policy)
  - 5 columns: #, Decision, Current Behavior, Assumption, Miltenyi's Answer
  - Header: bold white text on dark-blue (#1F4E79) background
  - Column widths A:6 / B:48 / C:45 / D:32 / E:32 (the D + E columns
    were 13 + 32 in the original but per-event audit rows need more
    room than policy rows; widened symmetrically to give Assumption
    space without crushing Miltenyi's Answer)
  - Wrap text on every body cell
  - Freeze pane at A2

Re-runnable. Overwrites the target file each time.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Style constants — kept in lock-step with the review-window file ──
HEADER_FILL = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
WRAP_LEFT_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_CENTER_TOP = Alignment(wrap_text=True, vertical="top", horizontal="center")

HEADERS = ["#", "Decision", "Current Behavior", "Assumption", "Miltenyi's Answer"]
COL_WIDTHS = {"A": 6.0, "B": 48.0, "C": 45.0, "D": 32.0, "E": 32.0}


# Per-row payload type. Body columns only — the # column auto-fills
# from the sheet prefix + position.
Row = tuple[str, str, str]  # (decision, current_behavior, assumption)


# =====================================================================
# SHEET CONTENT
# =====================================================================
#
# Each sheet starts with OPEN POLICY questions (cross-cutting decisions
# the stakeholders should settle first), then a PER-EVENT AUDIT (one
# row per notify() call site in that module). Numbering is
# <sheet_no>.<row_no>.

# ── Sheet 1: Goal Notifications ──────────────────────────────────────
GOAL_NOTIFICATIONS: list[Row] = [
    # Open policy
    (
        "Should goal-related notifications appear in the app bell, by "
        "email, or both?",
        "Today the bell always shows a new alert when something happens "
        "with a goal. Emails only go out for a few specific events (like "
        "mentor assignment changes and the welcome email). Goal events "
        "themselves are bell-only.",
        "Keep the bell for every event. Add email only for the events "
        "that block someone's work — when an employee submits a goal "
        "that needs the mentor's approval, or when the mentor sends it "
        "back for changes.",
    ),
    (
        "Should users be able to turn off goal-related emails?",
        "There is no off-switch today. If the company has email turned "
        "on at the system level, every goal email goes to everyone who "
        "is supposed to receive one.",
        "Add an 'email me about my goals' toggle in each user's Profile "
        "page. Bell notifications stay on for everyone — they're a quiet "
        "feed inside the app, not spam.",
    ),
    (
        "Should HR get a separate summary of goal activity, or only be "
        "alerted on big events?",
        "Today HR doesn't get any per-event alert about goals. They see "
        "a 'stalled goals' card on the dashboard that flags goals "
        "waiting too long for the mentor's approval.",
        "Keep HR out of per-event alerts (avoids noise). Optional: a "
        "weekly summary email showing approval velocity and the stalled "
        "count, if HR finds that useful.",
    ),
    # Per-event rows
    (
        "When an employee submits a goal for approval, who should be "
        "told and how?",
        "The mentor sees a bell notification: 'Bob Builder submitted a "
        "goal for your approval.' No email is sent.",
        "Keep the bell. Add email — this blocks the workflow until the "
        "mentor acts on it.",
    ),
    (
        "When a mentor approves a single goal, how should the employee "
        "be told?",
        "The employee sees a bell notification: 'Your goal was "
        "approved.' No email.",
        "Keep the bell. Email is optional — this is good news that can "
        "wait for the employee to open the app.",
    ),
    (
        "When a mentor sends a goal back for revisions, how should the "
        "employee be told?",
        "The employee sees a bell notification that includes the "
        "mentor's feedback. No email.",
        "Keep the bell. Add email — the employee needs to revise and "
        "resubmit, so this blocks their next step.",
    ),
    (
        "When a mentor approves several goals at once (bulk approve), "
        "should each owner get a separate notification per goal?",
        "Today every goal owner gets a separate bell alert for each of "
        "their goals that was approved — so one employee can see "
        "multiple alerts in a row.",
        "Roll up multiple approvals from the same mentor into a single "
        "summary: '3 of your goals were approved' instead of 3 separate "
        "alerts.",
    ),
    (
        "When a mentor uses the 'Notify' button on a goal to send a "
        "free-text message, how should it be delivered?",
        "The employee sees a bell notification with whatever the mentor "
        "typed. No email.",
        "Keep the bell. Add an option for the mentor to also send the "
        "message as an email when they want to.",
    ),
    (
        "When an employee submits their H1 or H2 self-review on a goal, "
        "who should be told?",
        "The mentor sees a bell notification: 'Bob Builder submitted an "
        "H1 self-review.' No email.",
        "Keep the bell. Email is optional — the mentor's review queue "
        "shows this when they next open the app.",
    ),
    (
        "When a mentor submits their H1 or H2 review on an employee's "
        "goal, how should the employee be told?",
        "The employee sees a bell notification: 'Your mentor submitted "
        "their H1 review.' No email.",
        "Keep the bell. No email needed — this is read-only feedback "
        "for the employee, not something they need to act on right away.",
    ),
    (
        "POTENTIAL: Should the system automatically remind mentors "
        "about goals that have been waiting for their approval too long?",
        "Today HR sees a 'stalled goals' chase list on the dashboard. "
        "Mentors get no automatic reminder; they have to open the app "
        "and look.",
        "Add an automatic daily reminder (bell + email) to mentors with "
        "any goals waiting 7+ days for their approval. The day threshold "
        "should be configurable.",
    ),
]

# ── Sheet 2: Annual Review Notifications ─────────────────────────────
ANNUAL_REVIEW_NOTIFICATIONS: list[Row] = [
    # Open policy
    (
        "When something happens with an annual review, should multiple "
        "people be notified (e.g. mentor + HR), or just the one person "
        "directly involved?",
        "Today each event has exactly one recipient. The mentor is told "
        "when the employee submits their self-appraisal; the employee is "
        "told when the mentor finishes their evaluation. HR is not "
        "alerted per event — they see the calibration grid in the app.",
        "Keep the single-recipient default. Add an optional company-wide "
        "setting that gives HR a copy whenever a review changes state "
        "(self-appraisal submitted, mentor evaluation done, final "
        "rating published).",
    ),
    (
        "Should the employee be told when HR publishes or updates their "
        "final 'management' rating?",
        "Yes — today the employee sees a bell notification: 'Your final "
        "rating is now available' on first publish, or 'Your final "
        "rating was updated' if HR changes it later. If HR re-saves the "
        "same value, no notification fires.",
        "Keep this. Whether the employee can actually see the rating "
        "number itself is governed by separate visibility settings — "
        "that's not changed by this notification policy.",
    ),
    (
        "Should HR be told when every mentor in the company has finished "
        "their evaluations, so they know calibration can begin?",
        "Today HR has to check the calibration grid manually to see if "
        "everyone has filed.",
        "Add a one-time bell + email to HR when all employee reviews "
        "have reached the 'pending management' stage: 'Calibration "
        "ready — every mentor has filed their evaluation.'",
    ),
    # Per-event rows
    (
        "When an employee submits their annual self-appraisal, who "
        "should be told and how?",
        "The mentor sees a bell notification: 'Bob Builder submitted "
        "their FY26-27 self-appraisal.' No email.",
        "Keep the bell. Add email — the annual review is a once-a-year "
        "event; mentors should be nudged out of the app to start their "
        "own evaluation.",
    ),
    (
        "When a mentor submits their annual evaluation of an employee, "
        "how should the employee be told?",
        "The employee sees a bell notification: 'Your mentor submitted "
        "their evaluation for FY26-27.' No email.",
        "Keep the bell. Add email — it's a year-end milestone the "
        "employee should know about.",
    ),
    (
        "When HR publishes or updates an employee's final rating, how "
        "should the employee be told?",
        "The employee sees a bell notification only. No email.",
        "Keep the bell. Add email — this is the year-end "
        "compensation-adjacent event; it warrants a notification "
        "outside the app.",
    ),
]

# ── Sheet 3: Project Review Notifications ────────────────────────────
PROJECT_REVIEW_NOTIFICATIONS: list[Row] = [
    # Open policy
    (
        "Should PMs be automatically reminded as the review cycle close "
        "approaches?",
        "No reminders today. The PM sees their pending-reviews queue in "
        "the app and has to remember to file. HR sees a completion "
        "card on the dashboard but doesn't actively nudge PMs.",
        "Add an automatic reminder (bell + email) sent 7 days before the "
        "cycle closes, to any PM with reviews still pending.",
    ),
    (
        "When an employee rolls off a project mid-cycle, should the PM "
        "be told their review window for that employee has shortened?",
        "No automatic message today. The assignment end-date silently "
        "shortens the PM's window without any heads-up.",
        "Add a bell alert to the PM: 'Bob rolled off ProjectX; complete "
        "his Q2 review by year-end or skip it.' Email is optional.",
    ),
    (
        "Should the secondary evaluator be told when the PM submits "
        "their primary review, prompting them to add their impact "
        "statement?",
        "Yes — today the secondary gets a bell notification when the "
        "PM submits. No email.",
        "Keep the bell. Add an email option for secondaries who don't "
        "open the app daily.",
    ),
    # Per-event rows
    (
        "When a PM submits a project review for one of their team "
        "members, how should the employee be told?",
        "The employee sees a bell notification: 'Your PM submitted a "
        "project review for ProjectX.' No email.",
        "Keep the bell. Add email — the employee should know that "
        "feedback has been filed about their work on the project.",
    ),
    (
        "When a secondary evaluator adds their impact statement, how "
        "should the employee be told?",
        "The employee sees a bell notification: 'A secondary evaluator "
        "added impact on your ProjectX review.' No email.",
        "Keep the bell. Email is optional.",
    ),
    (
        "When a PM goes back and edits a project review they already "
        "submitted, how should the employee be told?",
        "The employee sees a bell notification: 'Your PM updated their "
        "project review for ProjectX.' No email.",
        "Keep the bell. Send an email only if the rating actually "
        "changed by a step or more on the 1–5 scale — otherwise minor "
        "edits would spam the employee's inbox.",
    ),
    (
        "When a project is created with a secondary evaluator, how "
        "should that person be told?",
        "The secondary sees a bell notification with the project name. "
        "No email.",
        "Keep the bell. Add email — the secondary needs to know "
        "they're on a new project; they likely don't poll the app to "
        "find out.",
    ),
    (
        "When a project's secondary evaluator is reassigned, who should "
        "be told and how?",
        "The old secondary gets a bell notification ('You have been "
        "removed'), and the new secondary gets one ('You have been "
        "assigned'). Bell only — no email.",
        "Keep the bell for both. Add email to the NEW secondary so "
        "they know they're on the hook.",
    ),
    (
        "When a project is marked complete, who should be told?",
        "Bell notification to: the PM, every assigned employee, the "
        "secondary. No email.",
        "Keep the bell for everyone. Email is opt-in.",
    ),
    (
        "When someone is rolled off a project mid-stream, who should "
        "be told?",
        "Bell notification to the person who rolled off and to their "
        "PM, showing the end date. No email.",
        "Keep the bell. Email is optional.",
    ),
]

# ── Sheet 4: Admin & Account Notifications ───────────────────────────
ADMIN_ACCOUNT_NOTIFICATIONS: list[Row] = [
    # Open policy
    (
        "When HR Miltenyi creates a new user account, should HR MyOrg "
        "be told about it for audit purposes?",
        "No today. The new user account is recorded with its creation "
        "timestamp, but HR MyOrg is not actively notified.",
        "Optional daily summary email to HR MyOrg: 'N new users were "
        "created today by [HR Miltenyi name].' Not per-event.",
    ),
    (
        "Should a deactivated user receive any notification (knowing "
        "they can't log in anyway)?",
        "No. The system blocks any new notifications from being sent "
        "to a deactivated account, since the user can't log in to read "
        "them.",
        "Keep this. If HR later reactivates the user, a 'welcome back' "
        "notification fires automatically (already implemented).",
    ),
    (
        "Should an employee's mentor be told when the employee's "
        "profile changes (function, designation, etc.)?",
        "No. Only mentor-assignment changes themselves fire a "
        "notification — the rest of the profile is silent.",
        "Skip for now. Mentors with many mentees would get a flood of "
        "alerts every time HR updates a function or designation.",
    ),
    # Per-event rows
    (
        "When HR creates a new user account, how should the new user "
        "be told?",
        "The new user gets a welcome email with their temporary "
        "password, plus a bell notification on first login: 'Welcome to "
        "PMS — your account is ready.'",
        "Keep both. Critical onboarding event.",
    ),
    (
        "When an employee's mentor is assigned, changed, or removed, "
        "who should be told?",
        "Both the employee and the new mentor (if any) get a bell "
        "notification and an email describing the change.",
        "Keep both. The mentor relationship is foundational — the "
        "employee needs to know who reviews their work.",
    ),
    (
        "When HR reactivates a previously deactivated user, how should "
        "the user be told?",
        "Bell notification + email: 'Your account has been reactivated. "
        "You can sign in again.'",
        "Keep both.",
    ),
    (
        "When a user requests a password reset, how is the reset link "
        "delivered?",
        "Email only (the reset link is valid for 15 minutes). No bell "
        "notification — the user is locked out of the app at that "
        "moment.",
        "Keep email-only. The user can't see the bell when they're "
        "locked out.",
    ),
    (
        "When a user successfully changes their own password, should "
        "they get a confirmation?",
        "Today there's no notification on success. The Change Password "
        "modal just closes.",
        "Add a confirmation email — protects the user if someone else "
        "made the change. Optionally also a short on-screen banner on "
        "next login.",
    ),
    (
        "When HR forces a user to change their password on next login, "
        "should the user be told in advance?",
        "On their next login the user sees an on-screen banner that "
        "gates the app until they pick a new password. No email is "
        "sent beforehand.",
        "Add an email when HR triggers this: 'Your password was reset "
        "by HR. You'll be asked to choose a new one when you sign in.'",
    ),
]

# ── Sheet 5: Cross-cutting Policy ────────────────────────────────────
CROSS_CUTTING_POLICY: list[Row] = [
    (
        "By default, which kinds of events should go where — bell, "
        "email, or both?",
        "Today every event writes to the bell. Emails are opt-in per "
        "event — currently only mentor-assignment changes, the welcome "
        "email, password reset, and reactivation send email.",
        "Default policy:\n"
        "  • Workflow-blocking events (approval requested, changes "
        "requested) → bell + email\n"
        "  • Status updates (your review was submitted) → bell + "
        "email opt-in\n"
        "  • Read-only feedback (your H1 review is in) → bell only\n"
        "  • Security events (password reset, login from a new "
        "device) → email always",
    ),
    (
        "Should low-priority emails be batched into a daily or weekly "
        "digest?",
        "Every email fires the moment its event happens. There's no "
        "digest today.",
        "Add an opt-in 'send me a daily digest at 9am' switch in "
        "Profile, covering non-blocking events. Workflow-blocking "
        "emails still fire immediately, regardless of digest setting.",
    ),
    (
        "Where do users manage their own notification preferences?",
        "Today there's no preferences screen. Everyone gets the same "
        "notifications.",
        "Add a Profile → Notifications page with on/off toggles per "
        "category (Goals / Annual Reviews / Project Reviews / Admin) "
        "and per channel (bell on/off, email on/off, digest frequency).",
    ),
    (
        "Should HR be able to BCC themselves on every notification "
        "(observation / audit mode)?",
        "No. HR has no way to see what notifications have been sent "
        "across the company.",
        "Skip for now. Future: a Notification Audit Log page in Admin "
        "that lists every notification sent across the company, with "
        "filters by category, recipient, and date.",
    ),
    (
        "Should emails include an unsubscribe link in the footer?",
        "Today's emails don't have any unsubscribe footer.",
        "Add a 'manage email preferences' link in the footer that takes "
        "the user to Profile → Notifications. Not a hard unsubscribe — "
        "security emails (like password reset) always go out.",
    ),
    (
        "Should emails respect a per-user quiet window (e.g. don't email "
        "between 8pm and 7am local time)?",
        "Today emails fire immediately, regardless of time of day.",
        "Add quiet hours to the per-user preferences page. Bell "
        "notifications still fire (the user reads them when they want). "
        "Security emails ignore quiet hours.",
    ),
    (
        "POTENTIAL: Should the system integrate with Slack or Microsoft "
        "Teams?",
        "Not implemented today.",
        "Future scope. Highest value for mentor approval queues and HR "
        "reminders — those are the people most likely to live in "
        "Slack/Teams during the workday.",
    ),
    (
        "POTENTIAL: Should the system support mobile push notifications "
        "(mobile app or browser push)?",
        "Not implemented today. The app is web-only.",
        "Future scope. Lower priority — email covers most cases until "
        "there's a mobile presence to justify the build.",
    ),
    (
        "POTENTIAL: Should the system send SMS for security-critical "
        "events (password reset, account locked)?",
        "Not implemented today. Email is the only out-of-app channel.",
        "Future scope. Trade-off: SMS cost and carrier reliability vs. "
        "email's reach. Email plus two-factor / passkeys may be enough.",
    ),
    (
        "How long should bell notifications be kept?",
        "Every notification is kept forever in the database. The bell "
        "shows the most recent 20.",
        "Auto-archive read notifications older than 90 days — they're "
        "still kept for audit, just not shown in the bell anymore. "
        "Unread notifications never auto-archive.",
    ),
    (
        "If someone is deactivated, should their past notifications "
        "stay visible in other users' bells?",
        "Yes. New notifications can no longer be sent to deactivated "
        "users, but old notifications written BY a now-deactivated "
        "person stay in the recipients' bells.",
        "Keep this. The notification records a real past event — "
        "erasing it would erase history.",
    ),
    (
        "Where should HR go to see what notifications have been sent "
        "across the company?",
        "Today HR has no way to see this. They can only see what each "
        "user sees by logging in as them — which they can't do.",
        "Add an HR-only Notifications Audit page (in Admin Panel) that "
        "filters notifications by category, recipient, and date range. "
        "Read-only — no editing.",
    ),
]


SHEETS: list[tuple[str, int, list[Row]]] = [
    ("Goal Notifications", 1, GOAL_NOTIFICATIONS),
    ("Annual Review Notifications", 2, ANNUAL_REVIEW_NOTIFICATIONS),
    ("Project Review Notifications", 3, PROJECT_REVIEW_NOTIFICATIONS),
    ("Admin & Account Notifications", 4, ADMIN_ACCOUNT_NOTIFICATIONS),
    ("Cross-cutting Policy", 5, CROSS_CUTTING_POLICY),
]


# =====================================================================
# RENDERER
# =====================================================================

def build_sheet(wb: Workbook, name: str, sheet_no: int, rows: list[Row]) -> None:
    """Render one sheet with the standard header + body."""
    ws = wb.create_sheet(title=name)

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER_TOP

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Body
    for row_idx, (decision, current, assumption) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=f"{sheet_no}.{row_idx - 1}").alignment = WRAP_CENTER_TOP
        ws.cell(row=row_idx, column=2, value=decision).alignment = WRAP_LEFT_TOP
        ws.cell(row=row_idx, column=3, value=current).alignment = WRAP_LEFT_TOP
        ws.cell(row=row_idx, column=4, value=assumption).alignment = WRAP_LEFT_TOP
        # Column 5 left blank for Miltenyi's Answer
        ws.cell(row=row_idx, column=5, value="").alignment = WRAP_LEFT_TOP

    # Freeze header
    ws.freeze_panes = "A2"


def main() -> None:
    wb = Workbook()
    # openpyxl creates a default "Sheet" — remove before adding ours
    default = wb.active
    wb.remove(default)

    for name, sheet_no, rows in SHEETS:
        build_sheet(wb, name, sheet_no, rows)

    out = Path(__file__).resolve().parent.parent / "docs" / "meetings" / "2026-05-25-notifications-decision-matrix.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    total_rows = sum(len(r) for _, _, r in SHEETS)
    print(f"Wrote {out}")
    print(f"Sheets: {len(SHEETS)}  /  Total decision rows: {total_rows}")


if __name__ == "__main__":
    main()
