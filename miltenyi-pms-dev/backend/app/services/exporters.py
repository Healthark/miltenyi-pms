"""
exporters.py — HR_MyOrg Excel-export sheet builders.

Each `build_*_sheet` function takes an openpyxl Worksheet, a DB session,
the org_id of the caller, and an optional FY filter (set of 4-digit FY
start years; None = "all time"). It writes a header row plus one row
per record, and returns the number of data rows written (excluding the
header). The route layer wraps these into either a single-sheet
Workbook (per-tab buttons) or a 4-sheet Workbook (centralised page).

Design notes:
- Soft-deleted rows are skipped per entity (users.is_deleted, etc.) but
  the users sheet KEEPS deactivated employees with `Is Active = No` so
  HR can see ex-employee records — they're often what the audit is for.
- Long Text columns (review paragraphs, comments) are written verbatim,
  no truncation. Excel handles them fine; row height auto-expands when
  HR enables wrap-text.
- Cycle/FY extraction reuses the same logic as goal_routes._goal_fy_year
  (kept inline here to avoid an awkward cross-module import).
- Secondary evaluations on a project review can be 0..N; they're
  concatenated into a single "Name: text | Name: text" cell rather than
  exploded into multiple rows. Multi-row export was rejected as it would
  break "one row per assignment per cycle" — easier to read this way.
"""

from __future__ import annotations

from typing import Iterable, Optional

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session, joinedload

from app.core.cycle_utils import extract_fy_year as _extract_fy_year
from app.models.annual_review_models import AnnualReview
from app.models.goal_models import Goal
from app.models.goal_mentor_review_models import GoalMentorReview
from app.models.goal_self_review_models import GoalSelfReview
from app.models.project_models import Project, ProjectAssignment
from app.models.project_review_models import ProjectReview
from app.models.user_models import User


# ── Styling ──────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")  # mid blue
_HEADER_ALIGN = Alignment(vertical="center", horizontal="left")

# Internal row IDs are not surfaced to HR — every sheet leads with a
# dense 1-based "Sr. No." column instead.
_SR_NO_HEADER = "Sr. No."


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    """Write a styled header row and freeze it so HR can scroll the body."""
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.freeze_panes = "A2"


def _auto_size_columns(ws: Worksheet, max_width: int = 60) -> None:
    """Rough auto-size: each column gets sized to its widest cell, capped
    at `max_width` so paragraph cells don't blow the sheet sideways."""
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        widest = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > widest:
                widest = length
        # +2 padding, clamp.
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
            min(max_width, max(10, widest + 2))
        )


# ── FY extraction ─────────────────────────────────────────────────────
# `_extract_fy_year` is imported from `app.core.cycle_utils` so it stays
# in sync with the dashboard route's available-FY computation.


def _passes_fy_filter(fy_year: Optional[int], fy_filter: Optional[set[int]]) -> bool:
    """A row passes if (a) no filter is set, or (b) its FY matches.
    Rows with an unrecognisable cycle still pass when no filter is set —
    HR sees the messy data and can clean it up rather than silently
    losing rows."""
    if fy_filter is None:
        return True
    return fy_year is not None and fy_year in fy_filter


# ── Helpers ───────────────────────────────────────────────────────────

def _user_meta(user: Optional[User]) -> tuple[str, str, str, str]:
    """(full_name, email, function_name, designation_name) — empty strings
    for missing pieces. Centralised so every sheet renders user info the
    same way."""
    if user is None:
        return ("", "", "", "")
    function_name = user.function.name if user.function else ""
    designation_name = user.designation.name if user.designation else ""
    return (
        user.full_name or "",
        user.email or "",
        function_name,
        designation_name,
    )


def _find_review(
    reviews: Iterable, half: str
) -> Optional[object]:
    """First non-draft review for the given cycle_half, or None."""
    for r in reviews:
        if r.cycle_half == half and not getattr(r, "is_draft", False):
            return r
    return None


# ── Users sheet ───────────────────────────────────────────────────────

def build_users_sheet(
    ws: Worksheet,
    db: Session,
    org_id: int,
    exclude_roles: Optional[Iterable[str]] = None,
) -> int:
    """One row per user, including soft-deleted ones (with Is Active = No).
    FY filter doesn't apply — users sheet is a directory snapshot.

    `exclude_roles` is the role-scoping hook used by the HR_Miltenyi
    export: pass `{"Mentor", "HR_MyOrg"}` and those rows never make it
    into the workbook. Mentor names that appear as references on other
    sheets (e.g. a Staff member's mentor in the Mentor column) are not
    affected — only the directory rows themselves are filtered out."""
    ws.title = "Users"
    _write_header(
        ws,
        [
            _SR_NO_HEADER,
            "Full Name",
            "Email",
            "Employee Code",
            "Function",
            "Designation",
            "Mentor",
            "Phone",
            "Is Active",
            "Created At",
        ],
    )

    query = (
        db.query(User)
        .options(
            joinedload(User.function),
            joinedload(User.designation),
            joinedload(User.mentor),
        )
        .filter(User.org_id == org_id)
    )
    if exclude_roles:
        query = query.filter(User.role.notin_(list(exclude_roles)))
    users = query.order_by(User.full_name.asc()).all()

    row = 2
    for u in users:
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=u.full_name)
        ws.cell(row=row, column=3, value=u.email)
        ws.cell(row=row, column=4, value=u.employee_code)
        ws.cell(row=row, column=5, value=u.function.name if u.function else "")
        ws.cell(
            row=row, column=6, value=u.designation.name if u.designation else ""
        )
        ws.cell(row=row, column=7, value=u.mentor.full_name if u.mentor else "")
        ws.cell(row=row, column=8, value=u.phone or "")
        ws.cell(row=row, column=9, value="No" if u.is_deleted else "Yes")
        ws.cell(
            row=row,
            column=10,
            value=u.created_at.replace(tzinfo=None) if u.created_at else None,
        )
        row += 1

    _auto_size_columns(ws)
    return row - 2


# ── Annual goals sheet ───────────────────────────────────────────────

def build_goals_sheet(
    ws: Worksheet,
    db: Session,
    org_id: int,
    fy_filter: Optional[set[int]] = None,
    user_id_filter: Optional[int] = None,
) -> int:
    """One row per annual goal. H1/H2 self + mentor reviews are inlined
    as four extra text columns; draft self-reviews are skipped (mentee
    private state). When `user_id_filter` is set, only that user's
    goals are written (per-employee export)."""
    ws.title = "Annual Goals"
    _write_header(
        ws,
        [
            _SR_NO_HEADER,
            "Employee",
            "Email",
            "Function",
            "Designation",
            "Mentor",
            "FY",
            "Cycle Name",
            "Title",
            "Description",
            "Approval Status",
            "H1 Self Review",
            "H1 Mentor Review",
            "H2 Self Review",
            "H2 Mentor Review",
            "Created At",
            "Updated At",
        ],
    )

    goals_q = (
        db.query(Goal)
        .options(
            joinedload(Goal.owner).joinedload(User.function),
            joinedload(Goal.owner).joinedload(User.designation),
            joinedload(Goal.manager),
            joinedload(Goal.self_reviews),
            joinedload(Goal.mentor_reviews),
        )
        .filter(
            Goal.org_id == org_id,
            Goal.goal_type == "annual",
            # Drafts are private mentee work — never exposed via export.
            Goal.approval_status != "draft",
        )
    )
    if user_id_filter is not None:
        goals_q = goals_q.filter(Goal.user_id == user_id_filter)
    goals = goals_q.order_by(Goal.created_at.desc()).all()

    row = 2
    for g in goals:
        fy_year = _extract_fy_year(g.cycle_name)
        if not _passes_fy_filter(fy_year, fy_filter):
            continue

        full_name, email, func_name, desig_name = _user_meta(g.owner)
        manager_name = g.manager.full_name if g.manager else ""

        h1_self = _find_review(g.self_reviews, "H1")
        h2_self = _find_review(g.self_reviews, "H2")
        h1_mentor = _find_review(g.mentor_reviews, "H1")
        h2_mentor = _find_review(g.mentor_reviews, "H2")

        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=full_name)
        ws.cell(row=row, column=3, value=email)
        ws.cell(row=row, column=4, value=func_name)
        ws.cell(row=row, column=5, value=desig_name)
        ws.cell(row=row, column=6, value=manager_name)
        ws.cell(row=row, column=7, value=fy_year)
        ws.cell(row=row, column=8, value=g.cycle_name or "")
        ws.cell(row=row, column=9, value=g.title or "")
        ws.cell(row=row, column=10, value=g.description or "")
        ws.cell(row=row, column=11, value=g.approval_status)
        ws.cell(
            row=row, column=12, value=h1_self.self_overall_review if h1_self else ""
        )
        ws.cell(
            row=row,
            column=13,
            value=h1_mentor.mentor_overall_review if h1_mentor else "",
        )
        ws.cell(
            row=row, column=14, value=h2_self.self_overall_review if h2_self else ""
        )
        ws.cell(
            row=row,
            column=15,
            value=h2_mentor.mentor_overall_review if h2_mentor else "",
        )
        ws.cell(
            row=row,
            column=16,
            value=g.created_at.replace(tzinfo=None) if g.created_at else None,
        )
        ws.cell(
            row=row,
            column=17,
            value=g.updated_at.replace(tzinfo=None) if g.updated_at else None,
        )
        row += 1

    _auto_size_columns(ws)
    return row - 2


# ── Annual reviews sheet ──────────────────────────────────────────────

def build_annual_reviews_sheet(
    ws: Worksheet,
    db: Session,
    org_id: int,
    fy_filter: Optional[set[int]] = None,
    user_id_filter: Optional[int] = None,
) -> int:
    """One row per (user, FY). Drafts of mentor stage are not exposed —
    only the published mentor_overall_review / mentor_performance_rating
    cells make it out. `user_id_filter` narrows to a single employee."""
    ws.title = "Annual Reviews"
    _write_header(
        ws,
        [
            _SR_NO_HEADER,
            "Full Name",
            "Email",
            "Function",
            "Designation",
            "Mentor",
            "Cycle Name",
            "FY",
            "Status",
            "Self Rating",
            "Self Overall Review",
            "Mentor Rating",
            "Mentor Overall Review",
            "Management Rating",
            "Final Rating",
            "Created At",
            "Updated At",
        ],
    )

    reviews_q = (
        db.query(AnnualReview)
        .options(
            joinedload(AnnualReview.employee).joinedload(User.function),
            joinedload(AnnualReview.employee).joinedload(User.designation),
            joinedload(AnnualReview.mentor),
        )
        .filter(AnnualReview.org_id == org_id)
    )
    if user_id_filter is not None:
        reviews_q = reviews_q.filter(AnnualReview.user_id == user_id_filter)
    reviews = reviews_q.order_by(
        AnnualReview.cycle_name.desc(), AnnualReview.created_at.desc()
    ).all()

    row = 2
    for r in reviews:
        fy_year = _extract_fy_year(r.cycle_name)
        if not _passes_fy_filter(fy_year, fy_filter):
            continue

        full_name, email, func_name, desig_name = _user_meta(r.employee)
        mentor_name = r.mentor.full_name if r.mentor else ""

        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=full_name)
        ws.cell(row=row, column=3, value=email)
        ws.cell(row=row, column=4, value=func_name)
        ws.cell(row=row, column=5, value=desig_name)
        ws.cell(row=row, column=6, value=mentor_name)
        ws.cell(row=row, column=7, value=r.cycle_name or "")
        ws.cell(row=row, column=8, value=fy_year)
        ws.cell(row=row, column=9, value=r.status)
        ws.cell(row=row, column=10, value=r.self_performance_rating)
        ws.cell(row=row, column=11, value=r.self_overall_review or "")
        ws.cell(row=row, column=12, value=r.mentor_performance_rating)
        ws.cell(row=row, column=13, value=r.mentor_overall_review or "")
        ws.cell(row=row, column=14, value=r.management_performance_rating)
        # Synthesize from management ?? mentor when the stored column is NULL
        # but the row is officially published. Rows rated before
        # set_management_rating started persisting final_performance_rating
        # would otherwise export a blank Final column.
        final_rating = r.final_performance_rating
        if final_rating is None and r.final_rating_enabled:
            final_rating = (
                r.management_performance_rating
                if r.management_performance_rating is not None
                else r.mentor_performance_rating
            )
        ws.cell(row=row, column=15, value=final_rating)
        ws.cell(
            row=row,
            column=16,
            value=r.created_at.replace(tzinfo=None) if r.created_at else None,
        )
        ws.cell(
            row=row,
            column=17,
            value=r.updated_at.replace(tzinfo=None) if r.updated_at else None,
        )
        row += 1

    _auto_size_columns(ws)
    return row - 2


# ── Project reviews sheet ─────────────────────────────────────────────

def build_project_reviews_sheet(
    ws: Worksheet,
    db: Session,
    org_id: int,
    fy_filter: Optional[set[int]] = None,
    user_id_filter: Optional[int] = None,
) -> int:
    """One row per assignment × cycle. Secondary evaluations are
    concatenated as "Name: statement | Name: statement" rather than
    exploded across multiple rows. `user_id_filter` narrows to a
    single employee."""
    ws.title = "Project Reviews"
    _write_header(
        ws,
        [
            "Employee",
            "Email",
            "Function",
            "Designation",
            "Project Name",
            "Project Code",
            "PM",
            "Cycle",
            "FY",
            "Status",
            "Rating",
            "Task Execution",
            "Ownership",
            "Project Management",
            "Client Deliverables",
            "Communication",
            "Mentoring",
            "Competency & Skills",
            "Impact Statement",
            "Secondary Evaluations",
            "Created At",
            "Updated At",
        ],
    )

    reviews_q = (
        db.query(ProjectReview)
        .options(
            joinedload(ProjectReview.employee).joinedload(User.function),
            joinedload(ProjectReview.employee).joinedload(User.designation),
            joinedload(ProjectReview.project).joinedload(Project.pm),
            joinedload(ProjectReview.secondary_evaluations),
        )
        .filter(
            ProjectReview.org_id == org_id,
            ProjectReview.is_deleted.is_(False),
        )
    )
    if user_id_filter is not None:
        reviews_q = reviews_q.filter(ProjectReview.user_id == user_id_filter)
    reviews = reviews_q.order_by(
        ProjectReview.cycle.desc(), ProjectReview.created_at.desc()
    ).all()

    row = 2
    for r in reviews:
        fy_year = _extract_fy_year(r.cycle)
        if not _passes_fy_filter(fy_year, fy_filter):
            continue

        full_name, email, func_name, desig_name = _user_meta(r.employee)
        project = r.project
        project_name = project.name if project else ""
        project_code = project.project_code if project else ""
        pm_name = (
            project.pm.full_name if project and getattr(project, "pm", None) else ""
        )

        # Combine secondary evaluations into one cell. Skip drafts so HR
        # only sees submitted statements; preserves the order they were
        # written (joinedload preserves created_at order via the relation).
        secondary_parts = []
        for se in r.secondary_evaluations or []:
            if se.status != "submitted":
                continue
            evaluator_name = (
                se.evaluator.full_name if se.evaluator else "Unknown evaluator"
            )
            secondary_parts.append(
                f"{evaluator_name}: {se.impact_statement or ''}"
            )
        secondary_cell = " | ".join(secondary_parts)

        ws.cell(row=row, column=1, value=full_name)
        ws.cell(row=row, column=2, value=email)
        ws.cell(row=row, column=3, value=func_name)
        ws.cell(row=row, column=4, value=desig_name)
        ws.cell(row=row, column=5, value=project_name)
        ws.cell(row=row, column=6, value=project_code)
        ws.cell(row=row, column=7, value=pm_name)
        ws.cell(row=row, column=8, value=r.cycle or "")
        ws.cell(row=row, column=9, value=fy_year)
        ws.cell(row=row, column=10, value=r.status)
        ws.cell(row=row, column=11, value=r.performance_group)
        ws.cell(row=row, column=12, value=r.comment_task_execution or "")
        ws.cell(row=row, column=13, value=r.comment_ownership or "")
        ws.cell(row=row, column=14, value=r.comment_project_management or "")
        ws.cell(row=row, column=15, value=r.comment_client_deliverables or "")
        ws.cell(row=row, column=16, value=r.comment_communication or "")
        ws.cell(row=row, column=17, value=r.comment_mentoring or "")
        ws.cell(row=row, column=18, value=r.comment_competency_skills or "")
        ws.cell(row=row, column=19, value=r.impact_statement or "")
        ws.cell(row=row, column=20, value=secondary_cell)
        ws.cell(
            row=row,
            column=21,
            value=r.created_at.replace(tzinfo=None) if r.created_at else None,
        )
        ws.cell(
            row=row,
            column=22,
            value=r.updated_at.replace(tzinfo=None) if r.updated_at else None,
        )
        row += 1

    _auto_size_columns(ws)
    return row - 2


# ── Projects sheet ────────────────────────────────────────────────────

def build_projects_sheet(ws: Worksheet, db: Session, org_id: int) -> int:
    """One row per project (including completed; excluding hard-deleted).
    The Active Team Members column is the comma-separated full names of
    every assignment with `end_date IS NULL` — at PMS scale (handful per
    project) the cell stays readable. Total Assignments Ever counts every
    historical assignment (active + ended) for forensic context."""
    ws.title = "Projects"
    _write_header(
        ws,
        [
            _SR_NO_HEADER,
            "Code",
            "Name",
            "Description",
            "PM",
            "Secondary Evaluator",
            "Status",
            "Start Date",
            "Expected End Date",
            "Completed At",
            "Completed By",
            "Active Team Size",
            "Active Team Members",
            "Total Assignments Ever",
            "Created At",
        ],
    )

    projects = (
        db.query(Project)
        .options(
            joinedload(Project.pm),
            joinedload(Project.secondary_evaluator),
            joinedload(Project.completed_by),
            joinedload(Project.assignments).joinedload(ProjectAssignment.user),
        )
        .filter(
            Project.org_id == org_id,
            Project.is_deleted.is_(False),
        )
        .order_by(Project.created_at.desc())
        .all()
    )

    row = 2
    for p in projects:
        active_members = [
            a.user.full_name
            for a in p.assignments
            if a.end_date is None and a.user is not None
        ]
        active_members.sort(key=lambda n: n.lower())

        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=p.project_code)
        ws.cell(row=row, column=3, value=p.name)
        ws.cell(row=row, column=4, value=p.description or "")
        ws.cell(row=row, column=5, value=p.pm.full_name if p.pm else "")
        ws.cell(
            row=row,
            column=6,
            value=(
                p.secondary_evaluator.full_name if p.secondary_evaluator else ""
            ),
        )
        ws.cell(row=row, column=7, value=p.status)
        ws.cell(row=row, column=8, value=p.start_date)
        ws.cell(row=row, column=9, value=p.expected_end_date)
        ws.cell(
            row=row,
            column=10,
            value=p.completed_at.replace(tzinfo=None) if p.completed_at else None,
        )
        ws.cell(
            row=row,
            column=11,
            value=p.completed_by.full_name if p.completed_by else "",
        )
        ws.cell(row=row, column=12, value=len(active_members))
        ws.cell(row=row, column=13, value=", ".join(active_members))
        ws.cell(row=row, column=14, value=len(p.assignments))
        ws.cell(
            row=row,
            column=15,
            value=p.created_at.replace(tzinfo=None) if p.created_at else None,
        )
        row += 1

    _auto_size_columns(ws)
    return row - 2


# ── Per-employee: profile sheet (key-value layout) ────────────────────

def build_profile_sheet(ws: Worksheet, db: Session, user: User) -> int:
    """Single user's identity sheet — rendered vertically (Field | Value)
    because a one-row-wide layout reads badly for a single record. Returns
    the number of fields written (not "rows of data" in the usual sense,
    but the same shape the audit log expects)."""
    ws.title = "Profile"
    _write_header(ws, ["Field", "Value"])

    function_name = user.function.name if user.function else ""
    designation_name = user.designation.name if user.designation else ""
    mentor_name = user.mentor.full_name if user.mentor else ""

    fields: list[tuple[str, object]] = [
        ("Full Name", user.full_name),
        ("Email", user.email),
        ("Employee Code", user.employee_code),
        ("Role", user.role),
        ("Function", function_name),
        ("Designation", designation_name),
        ("Mentor", mentor_name),
        ("Phone", user.phone or ""),
        ("Is Active", "No" if user.is_deleted else "Yes"),
        (
            "Created At",
            user.created_at.replace(tzinfo=None) if user.created_at else None,
        ),
        (
            "Updated At",
            user.updated_at.replace(tzinfo=None) if user.updated_at else None,
        ),
    ]

    row = 2
    for label, value in fields:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    _auto_size_columns(ws)
    return len(fields)


# ── Per-employee: project assignments sheet ───────────────────────────

def build_project_assignments_sheet(
    ws: Worksheet,
    db: Session,
    org_id: int,
    user_id: int,
) -> int:
    """One row per assignment (active + ended) for the given user. The
    point of this sheet is the "when assigned / when off / why" history
    that the org-wide Project Reviews sheet doesn't carry."""
    ws.title = "Project Assignments"
    _write_header(
        ws,
        [
            _SR_NO_HEADER,
            "Project Name",
            "Project Code",
            "Project Status",
            "PM",
            "Role on Project",
            "Function",
            "Assigned Date",
            "End Date",
            "Ended By",
            "Currently Active",
            "Created At",
        ],
    )

    assignments = (
        db.query(ProjectAssignment)
        .options(
            joinedload(ProjectAssignment.project).joinedload(Project.pm),
            joinedload(ProjectAssignment.ended_by),
            joinedload(ProjectAssignment.function),
        )
        .filter(
            ProjectAssignment.org_id == org_id,
            ProjectAssignment.user_id == user_id,
        )
        .order_by(ProjectAssignment.assigned_date.desc().nullslast())
        .all()
    )

    row = 2
    for a in assignments:
        project = a.project
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=project.name if project else "")
        ws.cell(row=row, column=3, value=project.project_code if project else "")
        ws.cell(row=row, column=4, value=project.status if project else "")
        ws.cell(
            row=row,
            column=5,
            value=project.pm.full_name if project and project.pm else "",
        )
        ws.cell(row=row, column=6, value=a.assignment_role or "")
        ws.cell(row=row, column=7, value=a.function.name if a.function else "")
        ws.cell(row=row, column=8, value=a.assigned_date)
        ws.cell(row=row, column=9, value=a.end_date)
        ws.cell(
            row=row, column=10, value=a.ended_by.full_name if a.ended_by else ""
        )
        ws.cell(row=row, column=11, value="Yes" if a.end_date is None else "No")
        ws.cell(
            row=row,
            column=12,
            value=a.created_at.replace(tzinfo=None) if a.created_at else None,
        )
        row += 1

    _auto_size_columns(ws)
    return row - 2

