"""
export_routes.py — HR_MyOrg Excel-export endpoints.

Five GET endpoints return .xlsx StreamingResponses:
    /export/users.xlsx              — single-sheet
    /export/goals.xlsx              — single-sheet
    /export/annual-reviews.xlsx     — single-sheet
    /export/project-reviews.xlsx    — single-sheet
    /export/all.xlsx?fy=…           — 4-sheet workbook with optional
                                       FY filter (comma-separated start
                                       years, e.g. "2025,2026"). No `fy`
                                       param ⇒ "all time".

Per-tab buttons on the frontend hit the four single-sheet routes
without `fy` (always dump everything authorised). The centralised
"Exports" admin page hits the combined route with an FY picker.

Every successful download writes one row to `export_audit_logs` so the
"who exported what, when" trail is queryable later. The audit insert
runs BEFORE the response is streamed so the row is committed before
the bytes leave the server (a network error mid-stream still leaves a
useful audit record).
"""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Annotated, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.api.dependencies import CurrentUser
from app.api.routes.admin_routes import _require_hr_any, _require_hr_myorg
from app.core.database import get_db
from app.models.export_audit_log_models import ExportAuditLog
from app.models.system_settings_models import SystemSettings
from app.models.user_models import Role, User
from app.services.exporters import (
    build_annual_reviews_sheet,
    build_goals_sheet,
    build_profile_sheet,
    build_project_assignments_sheet,
    build_project_reviews_sheet,
    build_projects_sheet,
    build_users_sheet,
)

router = APIRouter()
DbSession = Annotated[Session, Depends(get_db)]

# Excel MIME type. Browsers honour it + the Content-Disposition filename
# to trigger a save dialog rather than rendering inline.
_XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Roles whose user records HR_Miltenyi must not see in exports: Healthark
# Mentors and the Healthark HR super-admins. Mirrors the same set used
# elsewhere (admin_routes._authorize_user_mutation, UsersTab's
# PROTECTED_ROLES) so the boundary is consistent across surfaces.
_HEALTHARK_EXPORT_HIDDEN_ROLES = frozenset(
    [Role.MENTOR.value, Role.HR_MYORG.value]
)


def _hidden_roles_for(current_user: User) -> Optional[frozenset[str]]:
    """Return the set of roles to filter out of user-listing exports for
    the given caller, or None when no scoping is needed. HR_Miltenyi is
    the only role that gets a non-None scope today."""
    if current_user.role == Role.HR_MILTENYI.value:
        return _HEALTHARK_EXPORT_HIDDEN_ROLES
    return None


def _fiscal_start_month(db: Session, org_id: int) -> int:
    """Return the org's `fiscal_start_month` from SystemSettings,
    defaulting to 4 (April) when the row hasn't been created yet.
    Centralised so every FY-aware export builder uses the same anchor.
    """
    settings = (
        db.query(SystemSettings)
        .filter(SystemSettings.org_id == org_id)
        .first()
    )
    if settings and settings.fiscal_start_month:
        return int(settings.fiscal_start_month)
    return 4


# ── Helpers ───────────────────────────────────────────────────────────

def _require_hr_miltenyi(current_user: User) -> None:
    """Raise 403 unless the caller is HR_Miltenyi.

    Miltenyi org has no "HR" Function/Department row, so authorization
    here keys off the user's role rather than a function lookup. Kept
    separate from `_require_hr_myorg` because the two HR roles have
    different export scopes (Miltenyi sees only users/projects/project
    reviews; the in-house HR gets the fuller annual-review workbook)."""
    if current_user.role != Role.HR_MILTENYI.value:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only the Miltenyi HR can access this resource.",
        )


def _parse_fy_filter(fy: Optional[str]) -> Optional[set[int]]:
    """Parse the `?fy=2025,2026` query param into a set of 4-digit years.

    Returns None when the param is absent or empty (= "all time"). Raises
    400 on malformed input rather than silently filtering to nothing,
    which would otherwise hide an HR mistake in the URL."""
    if not fy or not fy.strip():
        return None
    parts = [p.strip() for p in fy.split(",") if p.strip()]
    years: set[int] = set()
    for p in parts:
        if not p.isdigit() or len(p) != 4:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Invalid FY year '{p}'. Expected comma-separated "
                    "4-digit start years like 2025,2026."
                ),
            )
        years.add(int(p))
    return years or None


def _fy_scope_audit_value(fy_filter: Optional[set[int]]) -> Optional[str]:
    """Render the FY filter back into the audit row's `fy_scope` column.
    Null = "all time" (no filter applied)."""
    if not fy_filter:
        return None
    return ",".join(str(y) for y in sorted(fy_filter))


def _date_suffix() -> str:
    """`2026-05-11` style date stamp for filenames. ISO order so HR's
    download folder sorts chronologically."""
    return date.today().isoformat()


def _filename_for(kind: str, fy_filter: Optional[set[int]]) -> str:
    """`pms-{kind}-{date}.xlsx` with an optional `FY{years}` infix when
    the centralised page narrowed scope."""
    suffix = _date_suffix()
    if fy_filter:
        years = "-".join(str(y) for y in sorted(fy_filter))
        return f"pms-{kind}-FY{years}-{suffix}.xlsx"
    return f"pms-{kind}-{suffix}.xlsx"


def _workbook_to_response(
    wb: Workbook, filename: str
) -> StreamingResponse:
    """Serialise the workbook to an in-memory buffer and return a
    StreamingResponse with the right MIME + filename headers."""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type=_XLSX_MIME,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


def _log_export(
    db: Session,
    user_id: int,
    data_type: str,
    row_count: int,
    fy_filter: Optional[set[int]],
) -> None:
    """Insert one audit row and commit immediately. Caller has already
    validated authz and built the workbook in memory; we commit before
    streaming so a mid-flight network error doesn't lose the audit."""
    db.add(
        ExportAuditLog(
            user_id=user_id,
            data_type=data_type,
            row_count=row_count,
            fy_scope=_fy_scope_audit_value(fy_filter),
        )
    )
    db.commit()


# ── Per-tab single-sheet exports ──────────────────────────────────────

@router.get("/users.xlsx")
def export_users(db: DbSession, current_user: CurrentUser):
    """Full user directory snapshot (active + deactivated).

    Open to both HR roles — HR_Miltenyi uses this from the per-tab button
    in Admin → Users. The query is already org-scoped on
    `current_user.org_id`. HR_Miltenyi callers additionally have
    Healthark's Mentor and HR_MyOrg rows filtered out via
    `_hidden_roles_for` so the workbook matches what they see in the
    in-app table."""
    _require_hr_any(current_user)
    wb = Workbook()
    rows = build_users_sheet(
        wb.active,
        db,
        current_user.org_id,
        exclude_roles=_hidden_roles_for(current_user),
    )
    _log_export(db, current_user.id, "users", rows, None)
    return _workbook_to_response(wb, _filename_for("users", None))


@router.get("/goals.xlsx")
def export_goals(db: DbSession, current_user: CurrentUser):
    """Every annual goal with H1/H2 self + mentor reviews inlined."""
    _require_hr_myorg(current_user)
    wb = Workbook()
    rows = build_goals_sheet(wb.active, db, current_user.org_id)
    _log_export(db, current_user.id, "goals", rows, None)
    return _workbook_to_response(wb, _filename_for("annual-goals", None))


@router.get("/annual-reviews.xlsx")
def export_annual_reviews(db: DbSession, current_user: CurrentUser):
    """Every annual review (self / mentor / management) per user × FY."""
    _require_hr_myorg(current_user)
    wb = Workbook()
    rows = build_annual_reviews_sheet(wb.active, db, current_user.org_id)
    _log_export(db, current_user.id, "annual_reviews", rows, None)
    return _workbook_to_response(wb, _filename_for("annual-reviews", None))


@router.get("/project-reviews.xlsx")
def export_project_reviews(db: DbSession, current_user: CurrentUser):
    """Every project review (PM + secondary evaluators) per assignment × cycle.

    Open to both HR roles — project reviews are in Miltenyi HR's scope
    (the dashboard surfaces the ProjectReviewCompletionCard for them).
    Org-scoped on `current_user.org_id`, so HR_Miltenyi only ever sees
    their own org's reviews."""
    _require_hr_any(current_user)
    wb = Workbook()
    rows = build_project_reviews_sheet(wb.active, db, current_user.org_id)
    _log_export(db, current_user.id, "project_reviews", rows, None)
    return _workbook_to_response(wb, _filename_for("project-reviews", None))


@router.get("/projects.xlsx")
def export_projects(db: DbSession, current_user: CurrentUser):
    """Every project (active + completed; excluding hard-deleted) with PM,
    secondary evaluator, lifecycle dates, and active team roster.

    Open to both HR roles — HR_Miltenyi uses this from the per-tab button
    in Admin → Projects. Org-scoped on `current_user.org_id`, so
    HR_Miltenyi only ever sees their own org's projects."""
    _require_hr_any(current_user)
    wb = Workbook()
    rows = build_projects_sheet(wb.active, db, current_user.org_id)
    _log_export(db, current_user.id, "projects", rows, None)
    return _workbook_to_response(wb, _filename_for("projects", None))


# ── Centralised combined workbook ─────────────────────────────────────

@router.get("/all.xlsx")
def export_all(
    db: DbSession,
    current_user: CurrentUser,
    fy: Annotated[Optional[str], Query()] = None,
):
    """Combined 5-sheet workbook (Users / Annual Goals / Annual Reviews /
    Projects / Project Reviews). `fy` is a comma-separated list of
    4-digit start years (e.g. `?fy=2025,2026`); when set, every sheet is
    narrowed to rows whose lifecycle / cycle overlaps any selected FY.
    When the filter is empty, every sheet returns all-time data
    (preserves the original behavior)."""
    _require_hr_myorg(current_user)
    fy_filter = _parse_fy_filter(fy)
    fiscal_start_month = _fiscal_start_month(db, current_user.org_id)

    wb = Workbook()
    # Workbook() creates a default "Sheet" we'll repurpose as Users.
    users_ws = wb.active
    users_rows = build_users_sheet(
        users_ws,
        db,
        current_user.org_id,
        fy_filter=fy_filter,
        fiscal_start_month=fiscal_start_month,
    )

    goals_ws = wb.create_sheet("Annual Goals")
    goals_rows = build_goals_sheet(
        goals_ws, db, current_user.org_id, fy_filter
    )

    reviews_ws = wb.create_sheet("Annual Reviews")
    reviews_rows = build_annual_reviews_sheet(
        reviews_ws, db, current_user.org_id, fy_filter
    )

    projects_ws = wb.create_sheet("Projects")
    projects_rows = build_projects_sheet(
        projects_ws,
        db,
        current_user.org_id,
        fy_filter=fy_filter,
        fiscal_start_month=fiscal_start_month,
    )

    project_reviews_ws = wb.create_sheet("Project Reviews")
    project_reviews_rows = build_project_reviews_sheet(
        project_reviews_ws, db, current_user.org_id, fy_filter
    )

    total = (
        users_rows
        + goals_rows
        + reviews_rows
        + projects_rows
        + project_reviews_rows
    )
    _log_export(db, current_user.id, "combined", total, fy_filter)

    return _workbook_to_response(wb, _filename_for("workbook", fy_filter))


# ── Miltenyi HR combined workbook ─────────────────────────────────────

@router.get("/miltenyi.xlsx")
def export_miltenyi(
    db: DbSession,
    current_user: CurrentUser,
    fy: Annotated[Optional[str], Query()] = None,
):
    """Three-sheet workbook (Users / Projects / Project Reviews) scoped
    for Miltenyi HR.

    Annual goals and annual reviews are intentionally omitted — Miltenyi
    HR's scope excludes those flows (see HrDashboard.tsx and
    annual_review_routes.py: "Miltenyi HR has no business in annual
    reviews"), so leaving them out of the export keeps the workbook to
    only the data Miltenyi HR actually uses.

    `fy` is a comma-separated list of 4-digit start years (e.g.
    `?fy=2025,2026`); when set, every sheet narrows to rows whose
    lifecycle / cycle overlaps any selected FY. Empty filter returns
    all-time data.
    """
    _require_hr_miltenyi(current_user)
    fy_filter = _parse_fy_filter(fy)
    fiscal_start_month = _fiscal_start_month(db, current_user.org_id)

    wb = Workbook()
    users_ws = wb.active
    # Filter Healthark's Mentor + HR_MyOrg rows out of the Users sheet
    # so HR_Miltenyi's combined workbook never carries those records.
    users_rows = build_users_sheet(
        users_ws,
        db,
        current_user.org_id,
        exclude_roles=_hidden_roles_for(current_user),
        fy_filter=fy_filter,
        fiscal_start_month=fiscal_start_month,
    )

    projects_ws = wb.create_sheet("Projects")
    projects_rows = build_projects_sheet(
        projects_ws,
        db,
        current_user.org_id,
        fy_filter=fy_filter,
        fiscal_start_month=fiscal_start_month,
    )

    project_reviews_ws = wb.create_sheet("Project Reviews")
    project_reviews_rows = build_project_reviews_sheet(
        project_reviews_ws, db, current_user.org_id, fy_filter
    )

    total = users_rows + projects_rows + project_reviews_rows
    _log_export(db, current_user.id, "miltenyi_combined", total, fy_filter)

    return _workbook_to_response(
        wb, _filename_for("miltenyi-workbook", fy_filter)
    )


# Per-employee export is intentionally NOT exposed to HR_Miltenyi.
# Deep per-employee bundles (profile + assignments + project reviews)
# remain available to HR_MyOrg via /employee/{user_id}.xlsx below.


# ── Per-employee bundle ───────────────────────────────────────────────

@router.get("/employee/{user_id}.xlsx")
def export_employee(
    user_id: int,
    db: DbSession,
    current_user: CurrentUser,
    fy: Annotated[Optional[str], Query()] = None,
):
    """Single-employee deep-dive workbook with five sheets:

        - Profile               (key/value identity card)
        - Annual Goals          (this user's goals, H1/H2 reviews inline)
        - Annual Reviews        (this user's annual reviews per FY)
        - Project Assignments   (overlapping selected FYs when set)
        - Project Reviews       (every PM/secondary evaluation received)

    HR_MyOrg only. 404 when the user lives in a different org or no
    longer exists. Soft-deleted users are intentionally exportable so
    HR can still pull ex-employee records.

    `fy` mirrors the same comma-separated 4-digit start years used by
    /all.xlsx. When set, Annual Goals / Annual Reviews / Project
    Assignments / Project Reviews are narrowed; the Profile sheet is an
    FY-agnostic identity card and is always included.
    """
    _require_hr_myorg(current_user)

    target = (
        db.query(User)
        .filter(User.id == user_id, User.org_id == current_user.org_id)
        .first()
    )
    if target is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found.",
        )

    fy_filter = _parse_fy_filter(fy)
    fiscal_start_month = _fiscal_start_month(db, current_user.org_id)

    wb = Workbook()
    profile_ws = wb.active
    profile_rows = build_profile_sheet(profile_ws, db, target)

    goals_ws = wb.create_sheet("Annual Goals")
    goals_rows = build_goals_sheet(
        goals_ws, db, current_user.org_id, fy_filter, user_id_filter=user_id
    )

    reviews_ws = wb.create_sheet("Annual Reviews")
    reviews_rows = build_annual_reviews_sheet(
        reviews_ws, db, current_user.org_id, fy_filter, user_id_filter=user_id
    )

    assignments_ws = wb.create_sheet("Project Assignments")
    assignment_rows = build_project_assignments_sheet(
        assignments_ws,
        db,
        current_user.org_id,
        user_id,
        fy_filter=fy_filter,
        fiscal_start_month=fiscal_start_month,
    )

    project_reviews_ws = wb.create_sheet("Project Reviews")
    project_reviews_rows = build_project_reviews_sheet(
        project_reviews_ws,
        db,
        current_user.org_id,
        fy_filter,
        user_id_filter=user_id,
    )

    total = (
        profile_rows
        + goals_rows
        + reviews_rows
        + assignment_rows
        + project_reviews_rows
    )
    # Encode the target user in fy_scope so the audit ledger is queryable
    # later (e.g. "show every per-employee export for user 42"). The
    # column is freeform string up to 64 chars.
    _log_export_with_scope(
        db, current_user.id, "employee", total, f"user:{user_id}"
    )

    # Filename: pms-employee-{slug}-{date}.xlsx where slug is the
    # employee's full name lower-snake-cased so HR's download folder is
    # readable without renaming.
    slug = _slugify(target.full_name or f"user-{user_id}")
    filename = f"pms-employee-{slug}-{_date_suffix()}.xlsx"
    return _workbook_to_response(wb, filename)


def _slugify(name: str) -> str:
    """Lowercase + collapse whitespace/punct to single hyphens. Keeps
    filenames safe across operating systems and email clients."""
    import re as _re

    slug = _re.sub(r"[^A-Za-z0-9]+", "-", name).strip("-").lower()
    return slug or "user"


def _log_export_with_scope(
    db: Session,
    user_id: int,
    data_type: str,
    row_count: int,
    fy_scope: Optional[str],
) -> None:
    """Variant of `_log_export` that takes a pre-formatted fy_scope string
    instead of an FY-filter set. Used by the per-employee export to
    stash `user:{id}` in the same column."""
    db.add(
        ExportAuditLog(
            user_id=user_id,
            data_type=data_type,
            row_count=row_count,
            fy_scope=fy_scope,
        )
    )
    db.commit()
